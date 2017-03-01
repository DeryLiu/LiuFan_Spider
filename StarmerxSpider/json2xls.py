#!/usr/bin/env python
# coding=utf-8
import simplejson as json
from openpyxl import Workbook


def write_sort_xlsx(filename, xlsxname):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'name'
    ws['B1'] = 'url'
    ws['C1'] = 'brand'
    ws['D1'] = 'star'
    ws['E1'] = 'rank'
    ws['F1'] = 'other ranks'
    ws['G1'] = 'seller count'
    ws['H1'] = 'images'
    datas = []
    with open(filename) as f:
        for l in f:
            data = json.loads(l.strip())
            datas.append(data)
    datas = sorted(datas, key=lambda d: d['rank'])
    i = 2
    for data in datas:
        try:
            ws['A%s' % i] = data['name']
            ws['B%s' % i] = data['url']
            ws['C%s' % i] = data['brand']
            ws['D%s' % i] = data['star']
            ws['E%s' % i] = data['rank']
            ws['F%s' % i] = ';'.join('{path}: {rank}'.format(path=r['path'], rank=r['rank']) for r in data['other_ranks'])
            ws['G%s' % i] = data['sellers']
            ws['H%s' % i] = ';'.join(data['images'])
            i += 1
        except UnicodeEncodeError as e:
            print(e)
    wb.save(xlsxname)


if __name__ == '__main__':
    write_sort_xlsx('data.txt', 'data.xlsx')
