#!/usr/bin/env python
# encoding=utf-8
import re
import pymysql
from openpyxl import load_workbook


def extract_excel(filename, coloumn):
    reg = re.compile(r'DY_(\d+)-HD_\d+')
    wb = load_workbook(filename)
    ws = wb.active
    ids = set()
    for cell in ws[coloumn]:
        value = cell.value
        if value is not None:
            m = reg.search(cell.value)
            if m is not None:
                ids.add(m.group(1))
    out_filename = filename.split('.')[0] + '.txt'
    with open(out_filename, 'wt') as f:
        f.write('\n'.join(ids))


def read_image_urls_from_mysql(host='127.0.0.1', user='root', password='Tianhu2015$$', db='openDB', ids=None):
    conn = pymysql.connect(host=host, port=3306, user=user, password=password, db=db)
    curs = conn.cursor()
    sql = "select m.product_id, d.image from product_detail_4 as d, product_4 as m where m.product_id in (%s) AND d.product_id = m.id;"
    curs.execute(sql % ','.join(ids))
    data = curs.fetchall()
    curs.close()
    conn.close()
    return data


def read_products(host='127.0.0.1', user='root', password='Tianhu2015$$', db='openDB', ids=None):
    conn = pymysql.connect(host=host, port=3306, user=user, password=password, db=db)
    curs = conn.cursor()
    sql = "select id from product_4 where product_id in (%s);"
    curs.execute(sql % ','.join(ids))
    data = curs.fetchall()
    curs.close()
    conn.close()
    return data


def read_ids(filename):
    ids = []
    with open(filename) as f:
        for line in f:
            if line.endswith('\n'):
                line = line[:-1]
            ids.append('"%s"' % line)
    return ids


def write_file(data):
    ids = set()
    images = set()
    for d in data:
        ids.add(d[0])
        images.update(eval(d[1]))
    with open('ids.txt', 'wt') as f:
        f.write('\n'.join(ids))
    with open('images.txt', 'wt') as f:
        f.write('\n'.join(images))


if __name__ == '__main__':
    extract_excel(u'DY 需要修图上架表.xlsx', 'A')
    read_image_urls_from_mysql()
